import pandas as pd 
import os
from os.path import exists
#package below helps creating excel spreadsheets
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows



"""
This program finds profit of an organization that has entered their client rates in the client rates spreadsheet and  entered costs in the costs spreadsheet.

For the costs spreadsheet update or add any cost value along row 7. You can label the cost in row 6. 

For the cliet rates spreadsheet enter the client names under "Location" and hourly rate under "Hourly Rate".
"""

# get path 
path = os.path.normpath(os.getcwd())


#get client rates spreadsheet
rates_df = pd.read_excel(path + '/Client Rates.xlsx')
#get costs spreadsheet
monthly_costs_df = pd.read_excel(path + '/Costs.xlsx')

#getting monthly cost
monthly_costs_df = monthly_costs_df.iloc[4:6,:].reset_index().drop(columns=['Unnamed: 0','index'])
monthly_cost = sum(monthly_costs_df.iloc[1,:])

#check if a error file is already there from a previous run
if exists(path + '/Error File.xlsx'):
    #if it does exist I remove it
    os.remove(path + '/Error File.xlsx')

#Function creates a sheet that has details on billables - 1099 contract costs
def create_df_w_profit(rates_df,hourly_times_df):
    rates_lst = []
    client_rates_dict = dict()
    #create list of rates and dictionary of hourly rates for each client in client rates spreadhseet
    for i,element in enumerate(rates_df["Location"]):
            client_rates_dict[(element.lower()).strip()] = rates_df.loc[i,'Hourly Rate']
            rates_lst.append((element.lower()).strip())
    
    hourly_times_lst = []
    #get rid of last row
    hourly_times_df = hourly_times_df.drop([len(hourly_times_df)-1])
    rates_lst

    #get the the name of each client in timesheet spreadsheet and put them in alphabeltical order
    for element in hourly_times_df["location"].unique():
        hourly_times_lst.append((element.lower()).strip())
    hourly_times_lst = sorted(hourly_times_lst)
    times_rate_dict = dict()
    
    #make copy
    hourly_times_lst_copy = hourly_times_lst.copy()
    
    #checking to see if the client names in the client rates spreadsheet match the client names in the timesheet spreadhsheet
    for element in hourly_times_lst:
        for i, rate in enumerate(rates_lst):
            if (element == rate):
                #making a dictionrary of clients and their rates
                times_rate_dict[element] = client_rates_dict[rate]
                #if there names match remove that client from 
                hourly_times_lst_copy.remove(rate)
    #if there is a client in the client rates and not an exact match in the timesheet spreadsheet we need to indicate that there is a spelling error 
    if len(hourly_times_lst_copy) != 0:
        #return a dataframe with the misspelled client name
        return(pd.DataFrame(["Check Error excel sheet"]),0,hourly_times_lst_copy)

    #convert the hourly cost in the hourly times spreadsheet to a float 
    hourly_times_df['regular'] = hourly_times_df['regular'].astype(float)
    #Strip the "$" 
    for i in range(len(hourly_times_df)):
        hourly_times_df.loc[i,'cost'] = hourly_times_df.loc[i,'cost'].strip('$')
        hourly_times_df.loc[i,'rate'] = hourly_times_df.loc[i,'rate'].strip('$')
    # convert these columsn to floats
    hourly_times_df['cost'] = hourly_times_df['cost'].astype(float)
    hourly_times_df['total'] = hourly_times_df['total'].astype(float)
    hourly_times_df['overtime'] = hourly_times_df['overtime'].astype(float)
    hourly_times_df['special'] = hourly_times_df['special'].astype(float)
    #rename a colummn
    hourly_times_df.rename(columns = {'total':'total hours'}, inplace = True)
    # drop a column and sum groups
    df = hourly_times_df.groupby('location').sum().drop(columns = ['eid'])
    
    
    # get the indexs in a list
    indx = []
    for i in df.index:
        indx.append(i.lower())
    df.index = indx

    billables = []
    rate_col = []
    # get the client and rates in a list so its easier to use later
    for client in df.index:
        for element in times_rate_dict.keys():
            if element == client:
                rate_col.append(times_rate_dict[element])
                billables.append(round(df.loc[element,'total hours'] * times_rate_dict[client],2))
    #see if "rate to client" and "billable totals" already exists
    try:
        df['rate to client'] = rate_col       
        df['billable totals'] = billables
    #if they don't exist need to refer user to the error spreadsheet
    except:
        return(pd.DataFrame(["Check Error excel sheet"]),0,hourly_times_lst_copy)

    #drop the shift_title columns
    try:
        df = df.drop(columns = ['shift_title'])

        #getting totals
        totals_lst = ['','','',round(sum(df['total hours']),2),round(sum(df['cost']),2),'',round(sum(df['billable totals']),2)]

        #calculating total billables - 1099 costs
        cost_1099 = round(sum(df['cost']),2)
        billables = round(sum(df['billable totals']),2)
        total_earned = billables-cost_1099
        #setting up last line in spreadsheet
        sum_lst = [total_earned,'','','','','','']

        #getting breakdown of totals
        df.loc[len(df),:] = totals_lst
        #showing the total earned
        df.loc[len(df),:] = sum_lst
        
        #setting up the indexes of final dataframe
        new_index = []
        for i, ind in enumerate(df.index):
            if (i >=len(df)-2):
                break
            new_index.append(ind)
        new_index.append('Totals')
        new_index.append('Billables - 1099 cost:')
        df.index = new_index
        #return the dataframe with 
        return (df,total_earned,[])
    except:
        
        #getting totals
        totals_lst = ['','','',round(sum(df['total hours']),2),round(sum(df['cost']),2),'',round(sum(df['billable totals']),2)]

        #calculating total billables - 1099 costs
        cost_1099 = round(sum(df['cost']),2)
        billables = round(sum(df['billable totals']),2)
        total_earned = billables-cost_1099
        #setting up last line in spreadsheet
        sum_lst = [total_earned,'','','','','','']

        #getting breakdown of totals
        df.loc[len(df),:] = totals_lst
        #showing the total earned
        df.loc[len(df),:] = sum_lst
        
        #setting up the indexes of final dataframe
        new_index = []
        for i, ind in enumerate(df.index):
            if (i >=len(df)-2):
                break
            new_index.append(ind)
        new_index.append('Totals')
        new_index.append('Billables - 1099 cost:')
        df.index = new_index
        #return the dataframe with 
        return (df,total_earned,[])

# previous_path = os.path.normpath(os.getcwd() + os.sep + os.pardir)
filelist = os.listdir(path +'/Timesheet Reports')

total_error_lst = []
#opening new workbook in excel
wb = openpyxl.Workbook()
#rename sheet
wb['Sheet'].title = "1099 costs and billable Details"
wb.save(path +'/Total Profit.xlsx')
wb.close()
total_earned = 0
#for loop for multiple timesheets
for i in filelist:
    if i.endswith(".xlsx"):
        hourly_df = pd.read_excel(path +'/Timesheet Reports/'+i )
        #call function
        print(i)
        df, earned,error_lst = create_df_w_profit(rates_df,hourly_df)
        #checking if spelling error
        if len(error_lst) > 0:
            total_error_lst += error_lst
        #if no spelling error add the dataframe to the excel spreadsheet
        else:
            total_earned += earned
            openpyxl.load_workbook(path +'/Total Profit.xlsx')
            ws = wb.active
            for r in dataframe_to_rows(df, index=True, header=True):
                ws.append(r)
            wb.save(path +'/Total Profit.xlsx')
            wb.close()
        wb.close()
#create new worksheet
ws1 = wb.create_sheet('Total Profit.xlsx', 0)
#if there is a spelling error need to tell user to check the Error File
if len(total_error_lst) > 0:
    del wb['1099 costs and billable Details']
    ws1['A1'] = 'Please check excel sheet titled "Error File.xlsx"'
    wb.save(path +'/Total Profit.xlsx')
    wb.close()
    # shows potential place of error
    spell_check_df = pd.DataFrame([total_error_lst],index=["Find client/clients in the client rates file. Update the name to these/this name."])
    spell_check_df.to_excel(path +'/Error File.xlsx')
else:
    #create a new worksheet that shows Total profit
    ws1['A1'] = 'Total Profit'
    ws1['A2'] = total_earned - monthly_cost
    wb.save(path +'/Total Profit.xlsx')
    wb.close()
